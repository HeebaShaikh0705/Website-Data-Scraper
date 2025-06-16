import requests
from bs4 import BeautifulSoup
import dash
from dash import html, dcc, Input, Output, State
import flask
import os
from openpyxl import Workbook

# Initialize Dash app
app = dash.Dash(__name__, external_stylesheets=[
    "https://fonts.googleapis.com/css2?family=Roboto&display=swap"
])
server = app.server
app.title = "Web Data Scraper"

# Layout
app.layout = html.Div(
    style={
        "fontFamily": "Roboto, sans-serif",
        "padding": "20px",
        "backgroundColor": "#f9f9f9"
    },
    children=[
        html.Div([
            html.H2("🌐 Web Data Scraper", style={"textAlign": "center", "color": "#2c3e50"}),

            html.Div([
                dcc.Input(
                    id="input-url",
                    type="text",
                    placeholder="Enter full website URL (e.g. https://example.com)",
                    style={
                        "width": "70%",
                        "padding": "10px",
                        "fontSize": "16px",
                        "borderRadius": "5px",
                        "border": "1px solid #ccc",
                        "marginRight": "10px"
                    }
                ),
                html.Button(
                    "Scrape",
                    id="scrape-btn",
                    style={
                        "padding": "10px 20px",
                        "backgroundColor": "#3498db",
                        "color": "white",
                        "border": "none",
                        "borderRadius": "5px",
                        "cursor": "pointer",
                        "fontWeight": "bold"
                    }
                )
            ], style={"display": "flex", "justifyContent": "center", "marginBottom": "20px"}),

            html.Div([
                dcc.Checklist(
                    id="data-options",
                    options=[
                        {"label": "Title", "value": "title"},
                        {"label": "Paragraphs", "value": "paragraphs"},
                        {"label": "Images", "value": "images"},
                        {"label": "Links", "value": "links"},
                        {"label": "Tables", "value": "tables"},
                    ],
                    inline=True,
                    style={"textAlign": "center"}
                )
            ])
        ],
        style={
            "backgroundColor": "#fff",
            "padding": "30px",
            "borderRadius": "10px",
            "boxShadow": "0 4px 8px rgba(0,0,0,0.1)",
            "marginBottom": "30px"
        }),

        html.Div(id="output")
    ]
)

def get_soup(url):
    if not url.startswith("http"):
        url = "https://" + url
    r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
    return BeautifulSoup(r.text, "html.parser")

def extract_data(soup, url):
    title = soup.title.string.strip() if soup.title else "No Title"
    paragraphs = [p.get_text(strip=True) for p in soup.find_all("p") if p.get_text(strip=True)]
    images = [requests.compat.urljoin(url, img.get("src")) for img in soup.find_all("img") if img.get("src")]
    links = [requests.compat.urljoin(url, a.get("href")) for a in soup.find_all("a") if a.get("href")]

    tables = []
    tables_html = []
    for t in soup.find_all("table"):
        headers = [th.get_text(strip=True) for th in t.find_all("th")]
        rows = [[td.get_text(strip=True) for td in tr.find_all("td")] for tr in t.find_all("tr")[1:]]
        tables.append({"headers": headers, "rows": rows})
        tables_html.append(str(t))

    return title, paragraphs, images, links, tables, tables_html

def write_excel_file(title, paragraphs, images, links, tables, options):
    wb = Workbook()
    main_ws = wb.active
    main_ws.title = "Main Content"

    row = 1
    if "title" in options:
        main_ws.cell(row=row, column=1, value="Title")
        main_ws.cell(row=row, column=2, value=title)
        row += 2

    if "paragraphs" in options:
        main_ws.cell(row=row, column=1, value="Paragraphs")
        for para in paragraphs:
            row += 1
            main_ws.cell(row=row, column=2, value=para)
        row += 2

    if "images" in options:
        main_ws.cell(row=row, column=1, value="Images")
        for img in images:
            row += 1
            main_ws.cell(row=row, column=2, value=img)
        row += 2

    if "links" in options:
        main_ws.cell(row=row, column=1, value="Links")
        for link in links:
            row += 1
            main_ws.cell(row=row, column=2, value=link)

    if "tables" in options:
        for i, table in enumerate(tables, start=1):
            ws = wb.create_sheet(title=f"Table {i}")
            if table["headers"]:
                for col, header in enumerate(table["headers"], start=1):
                    ws.cell(row=1, column=col, value=header)
            for r_idx, row in enumerate(table["rows"], start=2):
                for c_idx, cell in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=cell)

    file_path = "scraped_data.xlsx"
    wb.save(file_path)
    return file_path

def card_style():
    return {
        "backgroundColor": "#fff",
        "padding": "20px",
        "borderRadius": "10px",
        "boxShadow": "0 4px 8px rgba(0,0,0,0.05)",
        "marginBottom": "20px"
    }

@app.callback(
    Output("output", "children"),
    Input("scrape-btn", "n_clicks"),
    State("input-url", "value"),
    State("data-options", "value"),
)
def scrape(n, url, options):
    if not n or not url:
        return ""

    try:
        soup = get_soup(url)
        title, paragraphs, images, links, tables, tables_html = extract_data(soup, url)
        file_path = write_excel_file(title, paragraphs, images, links, tables, options)

        preview = []

        if "title" in options:
            preview.append(html.Div([
                html.H4("Title", style={"color": "#2c3e50"}),
                html.P(title)
            ], style=card_style()))

        if "paragraphs" in options:
            preview.append(html.Div([
                html.H4("Paragraphs", style={"color": "#2c3e50"}),
                html.Ul([html.Li(p) for p in paragraphs[:10]] or ["No paragraphs found."])
            ], style=card_style()))

        if "images" in options:
            preview.append(html.Div([
                html.H4("Images", style={"color": "#2c3e50"}),
                html.Div(
                    [html.Img(src=img, style={"height": "80px", "margin": "5px"}) for img in images[:10]] or [html.P("No images found.")],
                    style={"display": "flex", "flexWrap": "wrap"}
                )
            ], style=card_style()))

        if "links" in options:
            preview.append(html.Div([
                html.H4("Links", style={"color": "#2c3e50"}),
                html.Ul([html.Li(html.A(href=link, children=link, target="_blank")) for link in links[:10]] or ["No links found."])
            ], style=card_style()))

        if "tables" in options:
            preview.append(html.Div([
                html.H4("Tables", style={"color": "#2c3e50"}),
                *[html.Iframe(srcDoc=table_html, style={
                    "width": "100%", "height": "200px", "border": "1px solid #ccc", "marginBottom": "15px"}) for table_html in tables_html[:3]]
            ], style=card_style()))

        preview.append(html.Div([
            html.H4("Download Excel File"),
            html.A("Click here to download", href="/download/scraped_data.xlsx", target="_blank", style={
                "textDecoration": "none",
                "color": "#2980b9",
                "fontWeight": "bold"
            })
        ], style=card_style()))

        return preview

    except Exception as e:
        return html.Div(f"Error: {e}", style={"color": "red", "textAlign": "center"})

@app.server.route("/download/<path:filename>")
def serve_files(filename):
    filepath = os.path.join(os.getcwd(), filename)
    if os.path.exists(filepath):
        return flask.send_file(filepath, as_attachment=True)
    return "File not found", 404

if __name__ == "__main__":
    app.run(debug=True, port=8051)
